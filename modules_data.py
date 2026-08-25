# modules_data.py
import csv
import re
from pathlib import Path

from openpyxl import load_workbook

from data.openstax_plan import MODULE_OPENSTAX_MAP, OPENSTAX_CHAPTERS
from data.canonical_texts import CANONICAL_TEXTS, MODULE_CANONICAL_MAP

MICRO_MODULES = [{'id': 1,
  'title': 'Economic Thought & Modeling',
  'short_desc': 'Scarcity, opportunity cost, models-as-maps, budget constraints, PPM, and '
                'comparative advantage.',
  'big_questions': '### Big Questions\n\n- How do economists think about the world around us and how do they turn everyday choices into quantifiable models?\n- How do scarcity, rational choice and marginal analysis shape economic thought?\n- How can the budget constraint, Production Possibilities Model, and comparative advantage explain constraints, opportunity cost, tradeoffs, and growth?',
  'tier1_definitions': '### Tier 1 — Math Support & Core Definitions\n'
                       '\n'
                       'Tier 1 is the focus for students who may have difficulty with mathematics coming into the course. By the end of Tier 1, students should be able to define:\n'
                       '\n'
                       '- **scarcity**\n'
                       '- **opportunity cost**\n'
                       '- **factors of production**\n'
                       '- **positive statement**\n'
                       '- **normative statement**\n'
                       '- **budget constraint**\n'
                       '- **PPM**\n'
                       '- **comparative advantage**',
  'tier2_solid': '### Tier 2 — General Education Standard (Assessment Tier)\n\nTier 2 is the general education target: graphs, tables, formulas, and algebraic reasoning for students taking the course as part of general education. No coding, calculus, or linear algebra is required. By the end of Tier 2, students should be able to:\n\n- **Part 1:** Understand how to use classroom resources in order to meet course expectations.\n- **Part 2:** Identify and explain three key principles in economic thought and their philosophical origins: scarcity, rational choice and marginal analysis.\n- **Part 3:** Use cross-referenced resources to explore key features of economics as an academic discipline and timelines of economic history.\n- **Part 4:** Develop the budget constraint in order to understand how economist model individual consumer decision-making under constraints of income - with a focus on scarcity.\n- **Part 5:** Develop the Production Possibilities Model to form a deeper understanding of how economist model individual producer decision-making - with a focus on opportunity cost and tradeoffs.\n- **Part 6:** Use the PPCs of two countries to explore comparative advantage conceptually and explain how marginal analysis and trade is beneficial to growth and production.',
  'tier3_extensions': '### Tier 3 — Quantitative Extension\n'
                      '\n'
                      'Tier 3 is a quantitative extension for deeper work: advanced algebra, '
                      'introductory calculus, linear algebra, coding, simulation, or more rigorous '
                      'mathematical proof. By the end of Tier 3, students can choose to:\n'
                      '\n'
                      '- Use advanced algebra to generalize budget constraints and PPMs with '
                      'parameters.\n'
                      '- Use spreadsheet or Python code to generate frontier graphs and '
                      'sensitivity checks.\n'
                      '- Compare linear and nonlinear opportunity-cost structures formally.\n'
                      '- Begin rigorous model critique by identifying assumptions as mathematical '
                      'restrictions.\n'
                      '- Formalize the **models-as-maps** critique by writing assumptions as '
                      'constraints and testing what changes when they are relaxed.',
  'materials': {'models': [{'label': 'Budget Constraint', 'url': '?model=Budget%20Constraint'},
                           {'label': 'PPC', 'url': '?model=PPC'},
                           {'label': 'Comparative Advantage',
                            'url': '?model=Comparative%20Advantage'}],
                'readings': [{'chapter': 1,
                              'label': 'OpenStax Ch 1: Welcome to Economics!',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/1-introduction'},
                             {'chapter': 2,
                              'label': 'OpenStax Ch 2: Choice in a World of Scarcity',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/2-introduction-to-choice-in-a-world-of-scarcity'}],
                'extensions': [],
                'labs': [],
                'khan': [{'label': 'Module 1 Quiz - Basic Economic Concepts (Quiz 3)',
                          'url': 'https://www.khanacademy.org/economics-finance-domain/ap-microeconomics/basic-economic-concepts/resource-allocation-and-economic-systems/quiz/basic-economic-concepts-quiz-1'},
                         {'label': 'Module 1 Quiz - Basic Economic Concepts (Quiz 1)',
                          'url': 'https://www.khanacademy.org/economics-finance-domain/ap-microeconomics/basic-economic-concepts/comparative-advantage-and-trade/quiz/basic-economic-concepts-quiz-2'},
                         {'label': 'Module 1 Quiz - Basic Economic Concepts (Quiz 2)',
                          'url': 'https://www.khanacademy.org/economics-finance-domain/ap-microeconomics/basic-economic-concepts/16/quiz/basic-economic-concepts-quiz-3'}],
                'videos': [],
                'audio': [],
                'primary_texts': [{'label': 'Adam Smith — Wealth of Nations (Book I, Ch.1: '
                                            'Division of Labour)',
                                   'url': 'https://www.econlib.org/library/Smith/smWN1.html#PartI',
                                   'era': '1776',
                                   'tradition': 'Classical',
                                   'note': 'Division of labour as the engine of productivity and '
                                           'market coordination.'},
                                  {'label': 'David Ricardo — Principles (Ch.1: Value)',
                                   'url': 'https://www.econlib.org/library/Ricardo/ricP1.html',
                                   'era': '1817',
                                   'tradition': 'Classical',
                                   'note': 'Labor-cost value and tradeoffs that underpin '
                                           'comparative advantage.'},
                                  {'label': 'J.S. Mill — On the Definition of Political Economy',
                                   'url': 'https://www.econlib.org/library/Mill/mlP.html',
                                   'era': '1844',
                                   'tradition': 'Classical',
                                   'note': "Defines political economy's scope and method as the "
                                           'science of incentives.'},
                                  {'label': 'Banerjee & Duflo — The Economic Lives of the Poor',
                                   'url': 'https://economics.mit.edu/files/4289',
                                   'era': '2007',
                                   'tradition': 'Development Micro',
                                   'note': 'Household decisions under poverty constraints using '
                                           'field evidence.'}],
                'slides': 'https://www.canva.com/design/DAGwfT91g1g/OdJml0o6MBrAhTj7OZbYCQ/view?utm_content=DAGwfT91g1g&utm_campaign=designshare&utm_medium=link2&utm_source=uniquelinks&utlId=h17b226aba2',
                'guided_notes': 'https://www.canva.com/design/DAGkcbYfGCc/ftT9TolnQqSJ47GTXdmhvQ/view?utm_content=DAGkcbYfGCc&utm_campaign=designshare&utm_medium=link2&utm_source=uniquelinks&utlId=he82f71711d'},
  'openstax': {'core': [{'chapter': 1,
                         'label': 'OpenStax Ch 1: Welcome to Economics!',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/1-introduction'},
                        {'chapter': 2,
                         'label': 'OpenStax Ch 2: Choice in a World of Scarcity',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/2-introduction-to-choice-in-a-world-of-scarcity'}],
               'optional': [{'chapter': 3,
                             'label': 'OpenStax Ch 3: Demand and Supply',
                             'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/3-introduction-to-demand-and-supply'}]},
  'format': 'IN_PERSON'},
 {'id': 2,
  'title': 'Choice',
  'short_desc': 'Labor-leisure feasible frontiers, intertemporal choice, behavioral limits, and '
                'institutional constraints.',
  'big_questions': '### Big Questions\n\n- How do utility, marginal utility and diminishing marginal utility influence human behavior?\n- How does a Feasible Set help explain individual labor choice and income flow?\n- How do limited time, cognition, and information influence choice?',
  'tier1_definitions': '### Tier 1 — Math Support & Core Definitions\n'
                       '\n'
                       'Tier 1 is the focus for students who may have difficulty with mathematics coming into the course. By the end of Tier 1, students should be able to define:\n'
                       '\n'
                       '- **production function**\n'
                       '- **marginal product**\n'
                       '- **feasible frontier**\n'
                       '- **MRT**\n'
                       '- **intertemporal budget constraint**\n'
                       '- **discount rate**\n'
                       '- **loss aversion**\n'
                       '- **bandwidth tax**\n'
                       '- **capabilities**',
  'tier2_solid': '### Tier 2 — General Education Standard (Assessment Tier)\n\nTier 2 is the general education target: graphs, tables, formulas, and algebraic reasoning for students taking the course as part of general education. No coding, calculus, or linear algebra is required. By the end of Tier 2, students should be able to:\n\n- **Part 1:** Use the asynchronous assignment to define utility, marginal utility and the influence of diminishing marginal utility on human behavior.\n- **Part 2:** Develop an individual production function and corresponding Feasible Set (tradeoff between work and free-time) - use this to understand individual labor choice and income flow.\n- **Part 3:** Review indifference curves and consumption bundles - use budget constraints and indifferences curves to graphically represent optimal choice in the market for goods.\n- **Part 4:** Use your understanding of optimal choice to develop the intertemporal choice model to graphically illustrate the lending and borrowing behavior of individuals based on preference - consider the difference between borrowing as a consumer against borrowing as a producer.\n- **Part 5:** Explore core behavioral economic scholarship to develop a working conceptual model of bounded rationality and how limited time, cognition, and information influence choice.',
  'tier3_extensions': '### Tier 3 — Quantitative Extension\n'
                      '\n'
                      'Tier 3 is a quantitative extension for deeper work: advanced algebra, '
                      'introductory calculus, linear algebra, coding, simulation, or more rigorous '
                      'mathematical proof. By the end of Tier 3, students can choose to:\n'
                      '\n'
                      '- Derive feasible-frontier equations symbolically and analyze parameter '
                      'changes with advanced algebra.\n'
                      '- Use introductory calculus to interpret marginal product, MRS, and '
                      'optimality conditions.\n'
                      '- Code labor-leisure and intertemporal simulations with sliders for wage '
                      'and interest-rate shocks.\n'
                      '- Explore present-biased choice with simple dynamic or recursive models.\n'
                      '- Code or derive extensions that compare rational intertemporal '
                      'optimization with present-biased behavior.',
  'materials': {'models': [{'label': 'Production Function and Marginal Product',
                            'url': '?model=Production%20Function%20and%20Marginal%20Product'},
                           {'label': 'Labor-Leisure Choice',
                            'url': '?model=Labor-Leisure%20Choice'},
                           {'label': 'Utility', 'url': '?model=Utility'},
                           {'label': 'Optimal Choice', 'url': '?model=Optimal%20Choice'},
                           {'label': 'Intertemporal Choice',
                            'url': '?model=Intertemporal%20Choice'}],
                'readings': [{'chapter': 2,
                              'label': 'OpenStax Ch 2: Choice in a World of Scarcity',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/2-introduction-to-choice-in-a-world-of-scarcity'},
                             {'chapter': 6,
                              'label': 'OpenStax Ch 6: Consumer Choices',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/6-introduction-to-consumer-choices'}],
                'extensions': [],
                'labs': [],
                'khan': [{'label': 'Module 6 Quiz - Basic Economic Concepts (Quiz 3)',
                          'url': 'https://www.khanacademy.org/economics-finance-domain/ap-microeconomics/basic-economic-concepts/16/quiz/basic-economic-concepts-quiz-3'}],
                'videos': [],
                'audio': [],
                'primary_texts': [{'label': 'Alfred Marshall — Principles (Book III: Demand, '
                                            'Utility, and Value)',
                                   'url': 'https://oll.libertyfund.org/title/marshall-principles-of-economics-8th-ed',
                                   'era': '1890',
                                   'tradition': 'Neoclassical',
                                   'note': 'Marginal utility and demand curves; consumer surplus '
                                           'framing.'},
                                  {'label': 'F.A. Hayek — The Use of Knowledge in Society',
                                   'url': 'https://www.econlib.org/library/Essays/hykKnw1.html',
                                   'era': '1945',
                                   'tradition': 'Austrian',
                                   'note': 'Prices coordinate dispersed knowledge; spontaneous '
                                           'order argument.'}],
                'slides': 'https://www.canva.com/design/DAGwf10bfYw/salIXy7kgdoFrE_JiMkAVg/view?utm_content=DAGwf10bfYw&utm_campaign=designshare&utm_medium=link2&utm_source=uniquelinks&utlId=hb547e94ffd',
                'guided_notes': 'https://www.canva.com/design/DAG5xP0dSpk/kJU_J-g6stn9rT3Q7qJCGw/view?utm_content=DAG5xP0dSpk&utm_campaign=designshare&utm_medium=link2&utm_source=uniquelinks&utlId=h3d387c98b9'},
  'openstax': {'core': [{'chapter': 2,
                         'label': 'OpenStax Ch 2: Choice in a World of Scarcity',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/2-introduction-to-choice-in-a-world-of-scarcity'},
                        {'chapter': 6,
                         'label': 'OpenStax Ch 6: Consumer Choices',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/6-introduction-to-consumer-choices'}],
               'optional': [{'chapter': 17,
                             'label': 'OpenStax Ch 17: Financial Markets',
                             'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/17-introduction-to-financial-markets'}]},
  'format': 'ASYNC'},
 {'id': 3,
  'title': 'Supply and Demand',
  'short_desc': 'Utility, indifference curves, optimal bundles, demand and supply, equilibrium, '
                'and market shifts.',
  'big_questions': '### Big Questions\n\n- How do budget constraints and indifferences curves graphically represent optimal choice in the market for goods?\n- What is the mathematical significance of where supply and demand intersect?\n- What can be determined about market price and market quantity when both the supply and demand curve shift at the same time?',
  'tier1_definitions': '### Tier 1 — Math Support & Core Definitions\n'
                       '\n'
                       'Tier 1 is the focus for students who may have difficulty with mathematics coming into the course. By the end of Tier 1, students should be able to define:\n'
                       '\n'
                       '- **utility**\n'
                       '- **marginal utility**\n'
                       '- **indifference curve**\n'
                       '- **MRS**\n'
                       '- **optimal bundle**\n'
                       '- **law of demand**\n'
                       '- **law of supply**\n'
                       '- **market equilibrium**\n'
                       '- **normal good**\n'
                       '- **inferior good**',
  'tier2_solid': '### Tier 2 — General Education Standard (Assessment Tier)\n\nTier 2 is the general education target: graphs, tables, formulas, and algebraic reasoning for students taking the course as part of general education. No coding, calculus, or linear algebra is required. By the end of Tier 2, students should be able to:\n\n- **Part 3:** Review indifference curves and consumption bundles - use budget constraints and indifferences curves to graphically represent optimal choice in the market for goods.\n- **Part 1:** Define the **law of demand -** determine graphical representations of demand and changes in **quantity demanded**.\n- **Part 2:** Define the **law of supply** - determine graphical representations of supply and changes in **quantity supplied.**\n- **Part 3:** Explore the mathematical significance of **where supply and demand intersect a**nd practice graphical determination of **market price and market quantity.**\n- **Part 4:** Review the demand curve and explore what **shifts demand left or right** on a coordinate plane**.**\n- **Part 5:** Review the supply curve and explore what **shifts supply left or right** on a coordinate plane**.**\n- **Part 6:** Isolate what can be determined about market price and market quantity when both the supply and demand curve shift at the same time.',
  'tier3_extensions': '### Tier 3 — Quantitative Extension\n'
                      '\n'
                      'Tier 3 is a quantitative extension for deeper work: advanced algebra, '
                      'introductory calculus, linear algebra, coding, simulation, or more rigorous '
                      'mathematical proof. By the end of Tier 3, students can choose to:\n'
                      '\n'
                      '- Use constrained optimization notation for utility maximization and the '
                      'MRS = price-ratio condition.\n'
                      '- Apply introductory calculus to marginal utility and demand sensitivity.\n'
                      '- Use matrix or linear-system methods to solve multiple market equilibria.\n'
                      '- Code demand, supply, and consumer-choice simulations from parameterized '
                      'functions.\n'
                      '- Derive consumer demand from utility maximization and compare the '
                      'derivation with behavioral deviations.',
  'materials': {'models': [{'label': 'Demand (schedule → line)',
                            'url': '?model=Demand%20(schedule%20%E2%86%92%20line)'},
                           {'label': 'Supply (schedule → line)',
                            'url': '?model=Supply%20(schedule%20%E2%86%92%20line)'},
                           {'label': 'Market Model (Supply & Demand)',
                            'url': '?model=Market%20Model%20(Supply%20%26%20Demand)'},
                           {'label': 'Single Shifts', 'url': '?model=Single%20Shifts'},
                           {'label': 'Double Shifts', 'url': '?model=Double%20Shifts'}],
                'readings': [{'chapter': 3,
                              'label': 'OpenStax Ch 3: Demand and Supply',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/3-introduction-to-demand-and-supply'},
                             {'chapter': 6,
                              'label': 'OpenStax Ch 6: Consumer Choices',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/6-introduction-to-consumer-choices'}],
                'extensions': [],
                'labs': [],
                'khan': [{'label': 'Module 2 Quiz - Supply & Demand (Quiz 1)',
                          'url': 'https://www.khanacademy.org/economics-finance-domain/ap-microeconomics/unit-2-supply-and-demnd/22/quiz/unit-2-supply-and-demnd-quiz-1'}],
                'videos': [],
                'audio': [],
                'primary_texts': [{'label': 'Alfred Marshall — Principles (Book III: Demand, '
                                            'Utility, and Value)',
                                   'url': 'https://oll.libertyfund.org/title/marshall-principles-of-economics-8th-ed',
                                   'era': '1890',
                                   'tradition': 'Neoclassical',
                                   'note': 'Marginal utility and demand curves; consumer surplus '
                                           'framing.'},
                                  {'label': 'Angrist & Pischke — Randomization & Causality '
                                            '(excerpt)',
                                   'url': 'https://economics.mit.edu/files/7504',
                                   'era': '2010',
                                   'tradition': 'Applied Micro/Methods',
                                   'note': 'Causal inference logic and when randomization '
                                           'identifies treatment effects.'}],
                'slides': 'https://www.canva.com/design/DAGu3QAnbyo/X4tBmSFNVLX_Wjotx8TuKg/view?utm_content=DAGu3QAnbyo&utm_campaign=designshare&utm_medium=link2&utm_source=uniquelinks&utlId=h17f367ac75',
                'guided_notes': 'https://www.canva.com/design/DAGxSH8hCHk/I2kj289MFKsKjWqFfBHeUA/view?utm_content=DAGxSH8hCHk&utm_campaign=designshare&utm_medium=link2&utm_source=uniquelinks&utlId=hb3c9693bc0'},
  'openstax': {'core': [{'chapter': 3,
                         'label': 'OpenStax Ch 3: Demand and Supply',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/3-introduction-to-demand-and-supply'},
                        {'chapter': 6,
                         'label': 'OpenStax Ch 6: Consumer Choices',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/6-introduction-to-consumer-choices'}],
               'optional': [{'chapter': 4,
                             'label': 'OpenStax Ch 4: Labor and Financial Markets',
                             'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/4-introduction-to-labor-and-financial-markets'}]},
  'format': 'IN_PERSON'},
 {'id': 4,
  'title': 'Market Analysis: Elasticity & Efficiency',
  'short_desc': 'Elasticity, total revenue, surplus, price controls, deadweight loss, and tax '
                'incidence.',
  'big_questions': '### Big Questions\n\n- How does the price elasticity of demand for a good influence a firm’s total revenue?\n- How can surplus be used as a measure of welfare within a market?\n- How can deadweight loss evaluate market inefficiencies when conditions for market equilibrium are not met?',
  'tier1_definitions': '### Tier 1 — Math Support & Core Definitions\n'
                       '\n'
                       'Tier 1 is the focus for students who may have difficulty with mathematics coming into the course. By the end of Tier 1, students should be able to define:\n'
                       '\n'
                       '- **elasticity**\n'
                       '- **total revenue**\n'
                       '- **income elasticity**\n'
                       '- **cross-price elasticity**\n'
                       '- **consumer surplus**\n'
                       '- **producer surplus**\n'
                       '- **total surplus**\n'
                       '- **price floor**\n'
                       '- **price ceiling**\n'
                       '- **deadweight loss**\n'
                       '- **tax incidence**',
  'tier2_solid': '### Tier 2 — General Education Standard (Assessment Tier)\n\nTier 2 is the general education target: graphs, tables, formulas, and algebraic reasoning for students taking the course as part of general education. No coding, calculus, or linear algebra is required. By the end of Tier 2, students should be able to:\n\n- **Part 1:** Define inelastic, elastic, unit elastic, and perfect elasticities using the formula for price elasticity of demand.\n- **Part 2:** Define total revenue and how the price elasticity of demand for a good influences a firm’s total revenue.\n- **Part 3:** Explore other applications of price elasticity of demand to make determinations between substitutes, complements, inferior and normal goods.\n- **Part 4:** Apply concepts of elasticity to producers and the law of supply.\n- **Part 5:** Articulate what market factors determine elasticity and explore how elasticity changes with growth.\n- **Part 1:** Review market price and market quantity and define the theory of market  equilibrium.\n- **Part 2:** Explore how to use surplus as a measure of welfare within a market.\n- **Part 3:** Understand the potential impact of government intervention within markets using price-floors and price-ceilings.\n- **Part 4:** Calculate deadweight loss to evaluate market inefficiencies when conditions for market equilibrium are not met.\n- **Part 5:** Understand the big picture of the market model with an online game.\n- **Part 6:** Extension of the Market Model: tax indices and market interactions between elasticity and surplus.',
  'tier3_extensions': '### Tier 3 — Quantitative Extension\n'
                      '\n'
                      'Tier 3 is a quantitative extension for deeper work: advanced algebra, '
                      'introductory calculus, linear algebra, coding, simulation, or more rigorous '
                      'mathematical proof. By the end of Tier 3, students can choose to:\n'
                      '\n'
                      '- Derive point elasticity and compare it with midpoint elasticity using '
                      'calculus intuition.\n'
                      '- Use symbolic algebra to derive tax incidence formulas from linear supply '
                      'and demand.\n'
                      '- Code welfare diagrams that compute CS, PS, TS, tax revenue, and DWL '
                      'automatically.\n'
                      '- Extend to nonlinear curves and compare area calculations with '
                      'integration.\n'
                      '- Use calculus, integration, or code to compare linear and nonlinear '
                      'welfare, elasticity, and incidence results.',
  'materials': {'models': [{'label': 'Price Elasticity of Demand',
                            'url': '?model=Price%20Elasticity%20of%20Demand'},
                           {'label': 'Elasticity and Total Revenue',
                            'url': '?model=Elasticity%20and%20Total%20Revenue'},
                           {'label': 'Price Elasticity of Supply',
                            'url': '?model=Price%20Elasticity%20of%20Supply'},
                           {'label': 'Surplus', 'url': '?model=Surplus'},
                           {'label': 'Government Intervention: Price Floor',
                            'url': '?model=Government%20Intervention:%20Price%20Floor'},
                           {'label': 'Government Intervention: Price Ceiling',
                            'url': '?model=Government%20Intervention:%20Price%20Ceiling'},
                           {'label': 'Deadweight Loss', 'url': '?model=Deadweight%20Loss'},
                           {'label': 'Tax Incidence', 'url': '?model=Tax%20Incidence'}],
                'readings': [{'chapter': 5,
                              'label': 'OpenStax Ch 5: Elasticity',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/5-introduction-to-elasticity'},
                             {'chapter': 3,
                              'label': 'OpenStax Ch 3: Demand and Supply',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/3-introduction-to-demand-and-supply'},
                             {'label': 'Module 4 Reading - Labor and Financial Markets',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/4-introduction-to-labor-and-financial-markets'}],
                'extensions': [],
                'labs': [],
                'khan': [{'label': 'Module 4 Quiz - Supply & Demand (Quiz 3)',
                          'url': 'https://www.khanacademy.org/economics-finance-domain/ap-microeconomics/unit-2-supply-and-demnd/27/quiz/unit-2-supply-and-demnd-quiz-3'},
                         {'label': 'Module 3 Quiz - Supply & Demand (Quiz 2)',
                          'url': 'https://www.khanacademy.org/economics-finance-domain/ap-microeconomics/unit-2-supply-and-demnd/25/quiz/unit-2-supply-and-demnd-quiz-2'},
                         {'label': 'Module 4 Quiz - Supply & Demand (Quiz 4)',
                          'url': 'https://www.khanacademy.org/economics-finance-domain/ap-microeconomics/unit-2-supply-and-demnd/29-international-trade-and-public-policy/quiz/unit-2-supply-and-demnd-quiz-4'}],
                'videos': [],
                'audio': [],
                'primary_texts': [{'label': 'A.C. Pigou — The Economics of Welfare (Externalities)',
                                   'url': 'https://socialsciences.mcmaster.ca/econ/ugcm/3ll3/pigou/welfare1.pdf',
                                   'era': '1920',
                                   'tradition': 'Welfare',
                                   'note': 'Externalities and corrective taxes/subsidies to align '
                                           'private and social cost.'},
                                  {'label': 'R.H. Coase — The Problem of Social Cost',
                                   'url': 'https://chicagounbound.uchicago.edu/cgi/viewcontent.cgi?article=1002&context=law_and_economics',
                                   'era': '1960',
                                   'tradition': 'Institutional',
                                   'note': 'Transaction costs and property rights; bargaining can '
                                           'solve externalities.'},
                                  {'label': 'Elinor Ostrom — Governing the Commons (Nobel Lecture)',
                                   'url': 'https://www.nobelprize.org/uploads/2018/06/ostrom_lecture.pdf',
                                   'era': '2009',
                                   'tradition': 'Institutional',
                                   'note': 'Commons self-governance design beyond '
                                           'privatize-or-regulate binaries.'},
                                  {'label': 'Jean Tirole — Market Power and Regulation (Nobel '
                                            'Lecture)',
                                   'url': 'https://www.nobelprize.org/uploads/2018/06/tirole-lecture.pdf',
                                   'era': '2014',
                                   'tradition': 'Industrial Organization',
                                   'note': 'Industrial organization theory for market power and '
                                           'regulation incentives.'}],
                'slides': 'https://www.canva.com/design/DAGwfq8bp5U/31Uju3ezcXK37ysvNjEstQ/view?utm_content=DAGwfq8bp5U&utm_campaign=designshare&utm_medium=link2&utm_source=uniquelinks&utlId=h2a359dfe37',
                'guided_notes': 'https://www.canva.com/design/DAGxT8SWQw4/fxTNVq45ccQYxK4R6RAlWg/view?utm_content=DAGxT8SWQw4&utm_campaign=designshare&utm_medium=link2&utm_source=uniquelinks&utlId=h2eb3f4c4cc'},
  'openstax': {'core': [{'chapter': 5,
                         'label': 'OpenStax Ch 5: Elasticity',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/5-introduction-to-elasticity'},
                        {'chapter': 3,
                         'label': 'OpenStax Ch 3: Demand and Supply',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/3-introduction-to-demand-and-supply'}],
               'optional': [{'chapter': 12,
                             'label': 'OpenStax Ch 12: Environmental Protection and Negative '
                                      'Externalities',
                             'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/12-introduction-to-environmental-protection-and-negative-externalities'},
                            {'chapter': 13,
                             'label': 'OpenStax Ch 13: Positive Externalities and Public Goods',
                             'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/13-introduction-to-positive-externalities-and-public-goods'},
                            {'label': 'Module 3 Reading - Consumer Choices (Optional)',
                             'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/6-introduction-to-consumer-choices'}]},
  'format': 'IN_PERSON'},
 {'id': 5,
  'title': 'Factor Markets',
  'short_desc': 'Product versus factor markets, VMP, land, labor, capital, rents, wages, interest, '
                'and monopsony preview.',
  'big_questions': '### Big Questions\n\n- How does Neoclassical economic theory explain income and resource distribution?\n- How is factor demand derived from the production decisions of firms?\n- How does marginal analysis help the firm derive their demand for labor?',
  'tier1_definitions': '### Tier 1 — Math Support & Core Definitions\n'
                       '\n'
                       'Tier 1 is the focus for students who may have difficulty with mathematics coming into the course. By the end of Tier 1, students should be able to define:\n'
                       '\n'
                       '- **factor market**\n'
                       '- **derived demand**\n'
                       '- **VMP**\n'
                       '- **MPL**\n'
                       '- **wage rate**\n'
                       '- **rental price**\n'
                       '- **loanable funds**\n'
                       '- **interest rate**\n'
                       '- **monopsony**',
  'tier2_solid': '### Tier 2 — General Education Standard (Assessment Tier)\n\nTier 2 is the general education target: graphs, tables, formulas, and algebraic reasoning for students taking the course as part of general education. No coding, calculus, or linear algebra is required. By the end of Tier 2, students should be able to:\n\n- **Part 1:** Explore Neoclassical economic theory as a theory of income and resource distribution.\n- **Part 2:** Identify the factors of production and begin to explore how their demand is derived from the production decisions of firms.\n- **Part 4:** Develop a static supply and demand model for capital markets (availability of credit and interest rates) - analyze the impact of price floors and ceilings on interest rates.\n- **Part 3:** Develop a static supply and demand model for land rents - analyze the impact of inelastic land supply throughout history.\n- **Part 5:** Develop a static supply and demand model for labor and explore market determents of wages.\n- **Part 6:** Explore what shifts the supply and demand of labor.\n- **Part 7:** Explore how the firm uses marginal analysis to derive their demand for labor.',
  'tier3_extensions': '### Tier 3 — Quantitative Extension\n'
                      '\n'
                      'Tier 3 is a quantitative extension for deeper work: advanced algebra, '
                      'introductory calculus, linear algebra, coding, simulation, or more rigorous '
                      'mathematical proof. By the end of Tier 3, students can choose to:\n'
                      '\n'
                      '- Derive VMP and MRP from production and revenue functions using '
                      'introductory calculus.\n'
                      '- Use systems of equations to model linked land, labor, and capital '
                      'markets.\n'
                      '- Code factor-market simulations with interdependent shocks.\n'
                      '- Use linear algebra notation to represent input combinations and '
                      'productivity changes.\n'
                      '- Model monopsony or linked factor markets with systems of equations or '
                      'matrix notation.',
  'materials': {'models': [{'label': 'Derived Demand and VMP',
                            'url': '?model=Derived%20Demand%20and%20VMP'},
                           {'label': 'Land + Rent', 'url': '?model=Land%20+%20Rent'},
                           {'label': 'Labor + Wage', 'url': '?model=Labor%20+%20Wage'},
                           {'label': 'Labor Market Policy',
                            'url': '?model=Labor%20Market%20Policy'},
                           {'label': 'Capital + Interest', 'url': '?model=Capital%20+%20Interest'}],
                'readings': [{'chapter': 4,
                              'label': 'OpenStax Ch 4: Labor and Financial Markets',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/4-introduction-to-labor-and-financial-markets'},
                             {'chapter': 14,
                              'label': 'OpenStax Ch 14: Labor Markets and Income',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/14-introduction-to-labor-markets-and-income'}],
                'extensions': [],
                'labs': [],
                'khan': [{'label': 'Module 5 Quiz - Factor Markets (Quiz 1)',
                          'url': 'https://www.khanacademy.org/economics-finance-domain/ap-microeconomics/factor-markets/ap-labor-marginal-product-rev/quiz/factor-markets-quiz-1'},
                         {'label': 'Module 5 Quiz - Basic Economic Concepts (Quiz 1)',
                          'url': 'https://www.khanacademy.org/economics-finance-domain/ap-microeconomics/basic-economic-concepts/resource-allocation-and-economic-systems/quiz/basic-economic-concepts-quiz-1'},
                         {'label': 'Module 5 Quiz - Production & Costs (Quiz 1)',
                          'url': 'https://www.khanacademy.org/economics-finance-domain/ap-microeconomics/production-cost-and-the-perfect-competition-model-temporary/33-long-run-p/quiz/production-cost-and-the-perfect-competition-model-temporary-quiz-1'}],
                'videos': [],
                'audio': [],
                'primary_texts': [{'label': 'David Ricardo — On Rent',
                                   'url': 'https://www.econlib.org/library/Ricardo/ricP5.html',
                                   'era': '1817',
                                   'tradition': 'Classical',
                                   'note': 'Differential rent and diminishing returns in '
                                           'agriculture and land use.'},
                                  {'label': 'Karl Marx — Wage Labour and Capital',
                                   'url': 'https://www.marxists.org/archive/marx/works/1847/wage-labour/',
                                   'era': '1847',
                                   'tradition': 'Classical Political Economy',
                                   'note': 'Explains wages, capital, and surplus '
                                           'value/exploitation.'},
                                  {'label': 'J.B. Clark — Distribution of Wealth (Marginal '
                                            'Productivity)',
                                   'url': 'https://oll.libertyfund.org/title/clark-the-distribution-of-wealth',
                                   'era': '1899',
                                   'tradition': 'Neoclassical',
                                   'note': 'Marginal productivity theory of factor payments.'},
                                  {'label': 'Autor, Dorn & Hanson — The China Syndrome',
                                   'url': 'https://economics.mit.edu/files/6613',
                                   'era': '2013',
                                   'tradition': 'Labor/Trade',
                                   'note': 'Trade shock from China and slow labor market '
                                           'adjustment costs.'}],
                'slides': 'https://www.canva.com/design/DAGwfzL7oXA/3XmNY9jjdA9aILtYOLcC5Q/view?utm_content=DAGwfzL7oXA&utm_campaign=designshare&utm_medium=link2&utm_source=uniquelinks&utlId=h49261ea0a9',
                'guided_notes': 'https://www.canva.com/design/DAGkcRAOK1Q/L294RJgfeF1FW2kLBPEr8A/view?utm_content=DAGkcRAOK1Q&utm_campaign=designshare&utm_medium=link2&utm_source=uniquelinks&utlId=ha3deda85e8'},
  'openstax': {'core': [{'chapter': 4,
                         'label': 'OpenStax Ch 4: Labor and Financial Markets',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/4-introduction-to-labor-and-financial-markets'},
                        {'chapter': 14,
                         'label': 'OpenStax Ch 14: Labor Markets and Income',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/14-introduction-to-labor-markets-and-income'}],
               'optional': [{'chapter': 7,
                             'label': 'OpenStax Ch 7: Production, Costs, and Industry Structure',
                             'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/7-introduction-to-production-costs-and-industry-structure'}]},
  'format': 'IN_PERSON'},
 {'id': 6,
  'title': 'Bridge: Markets, History & Global Economy',
  'short_desc': 'Course thesis synthesis, capitalism, great divergence, Malthus, externalities, '
                'and climate as market failure.',
  'big_questions': '### Big Questions\n\n- What explains economic inequality and divergence?\n- What role did the technological revolution play in growth?\n- What are the potential roles of capitalism, government, and the biosphere in economic growth?',
  'tier1_definitions': '### Tier 1 — Math Support & Core Definitions\n'
                       '\n'
                       'Tier 1 is the focus for students who may have difficulty with mathematics coming into the course. By the end of Tier 1, students should be able to define:\n'
                       '\n'
                       '- **capitalist revolution**\n'
                       '- **hockey stick**\n'
                       '- **great divergence**\n'
                       '- **Malthusian trap**\n'
                       '- **subsistence equilibrium**\n'
                       '- **negative externality**\n'
                       '- **social cost**\n'
                       '- **toolkit blind spots**',
  'tier2_solid': '### Tier 2 — General Education Standard (Assessment Tier)\n\nTier 2 is the general education target: graphs, tables, formulas, and algebraic reasoning for students taking the course as part of general education. No coding, calculus, or linear algebra is required. By the end of Tier 2, students should be able to:\n\n- **Objective 1:** Understand economic inequality and divergence.\n- **Objective 2**: Identify the technological revolution and its impact on growth.\n- **Objective 3**: Explore the potential roles of capitalism in economic growth.\n- **Objective 4**: Establish the importance of the government in capitalist economies.\n- **Objective 5**: Develop an understanding of the interaction between economics and the biosphere.',
  'tier3_extensions': '### Tier 3 — Quantitative Extension\n'
                      '\n'
                      'Tier 3 is a quantitative extension for deeper work: advanced algebra, '
                      'introductory calculus, linear algebra, coding, simulation, or more rigorous '
                      'mathematical proof. By the end of Tier 3, students can choose to:\n'
                      '\n'
                      '- Formalize the Malthusian model with equations for population, average '
                      'product, and subsistence.\n'
                      '- Use coded simulations to show poverty traps, external costs, or climate '
                      'damages over time.\n'
                      '- Connect the hockey-stick graph to growth-rate calculations and '
                      'exponential functions.\n'
                      '- Use rigorous mathematical critique to identify what the first-half '
                      'toolkit cannot represent.\n'
                      '- Simulate Malthusian dynamics, external costs, or climate damages to test '
                      'the bridge thesis rigorously.',
  'materials': {'models': [{'label': 'Malthusian Trap and Demographic Transition',
                            'url': '?model=Malthusian%20Trap%20and%20Demographic%20Transition'},
                           {'label': 'Three Engines and the Great Divergence',
                            'url': '?model=Three%20Engines%20and%20the%20Great%20Divergence'},
                           {'label': 'Poverty, GDP, and the Kuznets Curve',
                            'url': '?model=Poverty,%20GDP,%20and%20the%20Kuznets%20Curve'},
                           {'label': 'Climate Externality and the Atmosphere Commons',
                            'url': '?model=Climate%20Externality%20and%20the%20Atmosphere%20Commons'}],
                'readings': [],
                'extensions': [],
                'labs': [],
                'khan': [{'label': 'Module 7 Quiz - Basic Economic Concepts (Quiz 1)',
                          'url': 'https://www.khanacademy.org/economics-finance-domain/ap-microeconomics/basic-economic-concepts/resource-allocation-and-economic-systems/quiz/basic-economic-concepts-quiz-1'}],
                'videos': [],
                'audio': [],
                'primary_texts': [{'label': 'Kahneman & Tversky — Prospect Theory',
                                   'url': 'https://www.princeton.edu/~kahneman/docs/Publications/prospect_theory.pdf',
                                   'era': '1979',
                                   'tradition': 'Behavioral',
                                   'note': 'Prospect theory with loss aversion and reference '
                                           'dependence.'},
                                  {'label': 'Amartya Sen — Equality of What?',
                                   'url': 'https://dash.harvard.edu/bitstream/handle/1/9407592/EqualityofWhat.pdf',
                                   'era': '1979',
                                   'tradition': 'Welfare/Capabilities',
                                   'note': 'Capabilities approach to welfare beyond income; '
                                           'inequality as deprivation.'}]},
  'openstax': {'optional': [{'chapter': 19,
                             'label': 'OpenStax Ch 19: International Trade',
                             'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/19-introduction-to-international-trade'},
                            {'chapter': 20,
                             'label': 'OpenStax Ch 20: Globalization and Protectionism',
                             'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/20-introduction-to-globalization-and-protectionism'},
                            {'label': 'Module 7 Reading - Public Economy (Optional)',
                             'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/18-introduction-to-public-economy'}]},
  'format': 'BRIDGE'},
 {'id': 7,
  'title': 'Structural Inequality: Core + Game Theory Preview',
  'short_desc': 'Poverty traps, Lorenz curves, factor-market inequality, principal-agent problems, '
                'and informal game theory preview.',
  'big_questions': '### Big Questions\n\n- What is the difference between inequality of wealth vs. inequality of income?\n- How do limitations to the intertemporal choice model connect to the Poverty Trap?\n- How do Principal-Agent Problems influence lenders and borrowers?',
  'tier1_definitions': '### Tier 1 — Math Support & Core Definitions\n'
                       '\n'
                       'Tier 1 is the focus for students who may have difficulty with mathematics coming into the course. By the end of Tier 1, students should be able to define:\n'
                       '\n'
                       '- **Lorenz curve**\n'
                       '- **Gini coefficient**\n'
                       '- **poverty trap**\n'
                       '- **bandwidth tax**\n'
                       '- **MRP**\n'
                       '- **statistical discrimination**\n'
                       '- **monopsony**\n'
                       '- **adverse selection**\n'
                       '- **moral hazard**\n'
                       '- **strategic interdependence**',
  'tier2_solid': '### Tier 2 — General Education Standard (Assessment Tier)\n\nTier 2 is the general education target: graphs, tables, formulas, and algebraic reasoning for students taking the course as part of general education. No coding, calculus, or linear algebra is required. By the end of Tier 2, students should be able to:\n\n- **Part 1:** Develop a personal definition of inequality and understand the difference between inequality of wealth vs. inequality of income.\n- Part 3: Understand limitations to the intertemporal choice model - connect these limitations to  discourse around **The Poverty Trap** in current public policy literature.\n- **Part 4:** Develop an understanding of the Principal-Agent Problem and how this influences lenders and borrowers - in particular - highlight, how capital markets become a new landscape for inequality in Capitalist systems.',
  'tier3_extensions': '### Tier 3 — Quantitative Extension\n'
                      '\n'
                      'Tier 3 is a quantitative extension for deeper work: advanced algebra, '
                      'introductory calculus, linear algebra, coding, simulation, or more rigorous '
                      'mathematical proof. By the end of Tier 3, students can choose to:\n'
                      '\n'
                      '- Calculate Gini coefficients from grouped or microdata using formulas or '
                      'code.\n'
                      '- Use matrix payoff representations to prepare for formal game theory.\n'
                      '- Model credit exclusion with parameterized borrowing constraints and '
                      'interest-rate spreads.\n'
                      '- Use coding or data analysis to build Lorenz curves from real datasets.\n'
                      '- Use data or code to construct Lorenz curves, Gini coefficients, and '
                      'credit-constraint simulations.',
  'materials': {'models': [{'label': 'Structural Inequality Model',
                            'url': '?model=Structural%20Inequality%20Model'},
                           {'label': 'Lorenz Curve and Gini Coefficient',
                            'url': '?model=Lorenz%20Curve%20and%20Gini%20Coefficient'},
                           {'label': 'Credit Exclusion and Labor Power',
                            'url': '?model=Credit%20Exclusion%20and%20Labor%20Power'},
                           {'label': 'Game Theory Preview',
                            'url': '?model=Game%20Theory%20Preview'}],
                'readings': [{'chapter': 15,
                              'label': 'OpenStax Ch 15: Poverty and Economic Inequality',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/15-introduction-to-poverty-and-economic-inequality'},
                             {'chapter': 16,
                              'label': 'OpenStax Ch 16: Information, Risk, and Insurance',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/16-introduction-to-information-risk-and-insurance'},
                             {'label': 'Module 8 Reading - Labor Markets and Income',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/14-introduction-to-labor-markets-and-income'}],
                'extensions': [],
                'labs': [],
                'khan': [{'label': 'Module 8 Quiz - Market Failure (Quiz 2)',
                          'url': 'https://www.khanacademy.org/economics-finance-domain/ap-microeconomics/ap-consumer-producer-surplus/inequality/quiz/ap-consumer-producer-surplus-quiz-2'}],
                'videos': [],
                'audio': [],
                'primary_texts': [{'label': 'Joseph Schumpeter — Creative Destruction',
                                   'url': 'https://www.hup.harvard.edu/wp-content/uploads/2019/03/9780691165020-excerpt.pdf',
                                   'era': '1942',
                                   'tradition': 'Evolutionary',
                                   'note': 'Innovation-driven competition and cycles of creative '
                                           'destruction.'},
                                  {'label': 'Adam Smith — Wealth of Nations (Book I, Ch.1: '
                                            'Division of Labour)',
                                   'url': 'https://www.econlib.org/library/Smith/smWN1.html#PartI',
                                   'era': '1776',
                                   'tradition': 'Classical',
                                   'note': 'Division of labour as the engine of productivity and '
                                           'market coordination.'},
                                  {'label': 'Acemoglu, Johnson & Robinson — Institutions as the '
                                            'Fundamental Cause',
                                   'url': 'https://economics.mit.edu/files/4469',
                                   'era': '2005',
                                   'tradition': 'Political Economy',
                                   'note': 'Institutions shape incentives and long-run development '
                                           'paths.'}]},
  'openstax': {'core': [{'chapter': 15,
                         'label': 'OpenStax Ch 15: Poverty and Economic Inequality',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/15-introduction-to-poverty-and-economic-inequality'},
                        {'chapter': 16,
                         'label': 'OpenStax Ch 16: Information, Risk, and Insurance',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/16-introduction-to-information-risk-and-insurance'}],
               'optional': [{'chapter': 18,
                             'label': 'OpenStax Ch 18: Public Economy',
                             'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/18-introduction-to-public-economy'}]},
  'format': 'ASYNC'},
 {'id': 8,
  'title': 'Structural Inequality: Extensions',
  'short_desc': 'Technology diffusion, digital divide, AI bias, and climate as distributional '
                'injustice.',
  'big_questions': '### Big Questions\n\n- How do historic inequality within the global south and current dynamics of global income inequality connect?\n- Why doesn’t racial discrimination in capitalist systems quiet make rational or mathematical sense within traditional neoclassical thought?\n- What tradeoffs exist between production growth and fairness/equity?',
  'tier1_definitions': '### Tier 1 — Math Support & Core Definitions\n'
                       '\n'
                       'Tier 1 is the focus for students who may have difficulty with mathematics coming into the course. By the end of Tier 1, students should be able to define:\n'
                       '\n'
                       '- **S-curve diffusion**\n'
                       '- **network effects**\n'
                       '- **digital divide**\n'
                       '- **AI bias**\n'
                       '- **proxy bias**\n'
                       '- **fairness impossibility theorem**\n'
                       '- **climate risk exposure**',
  'tier2_solid': '### Tier 2 — General Education Standard (Assessment Tier)\n\nTier 2 is the general education target: graphs, tables, formulas, and algebraic reasoning for students taking the course as part of general education. No coding, calculus, or linear algebra is required. By the end of Tier 2, students should be able to:\n\n- **Part 2:** Understand historic inequality within the global south, while exploring current dynamics of global income inequality.\n- **Part 5:** Explore how racial discrimination in capitalist systems doesn’t quiet make rational or mathematical sense within traditional neoclassical thought.\n- **Part 6:** Explore how technology and education act as mitigating factors to the malthusian trap (production decline) AND inequality in a dynamic economy - explore potential tradeoffs between production growth and fairness/equity.\n- **Part 7:** Develop an original microeconomic thesis statement on inequality within the global economy.',
  'tier3_extensions': '### Tier 3 — Quantitative Extension\n'
                      '\n'
                      'Tier 3 is a quantitative extension for deeper work: advanced algebra, '
                      'introductory calculus, linear algebra, coding, simulation, or more rigorous '
                      'mathematical proof. By the end of Tier 3, students can choose to:\n'
                      '\n'
                      '- Fit or simulate S-curve diffusion models with logistic functions.\n'
                      '- Use confusion matrices to analyze algorithmic fairness metrics and '
                      'tradeoffs.\n'
                      '- Code simple classification or screening examples to show proxy '
                      'discrimination.\n'
                      '- Model climate exposure and factor-productivity impacts with indexed data '
                      'or matrix notation.\n'
                      '- Use confusion matrices, logistic curves, or coded simulations to evaluate '
                      'AI fairness and diffusion claims.',
  'materials': {'models': [{'label': 'Technology, AI Bias, and Climate Inequality',
                            'url': '?model=Technology,%20AI%20Bias,%20and%20Climate%20Inequality'},
                           {'label': 'Climate as Distributional Injustice',
                            'url': '?model=Climate%20as%20Distributional%20Injustice'},
                           {'label': 'AI Bias and Algorithmic Fairness',
                            'url': '?model=AI%20Bias%20and%20Algorithmic%20Fairness'}],
                'readings': [{'chapter': 15,
                              'label': 'OpenStax Ch 15: Poverty and Economic Inequality',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/15-introduction-to-poverty-and-economic-inequality'}],
                'extensions': [],
                'labs': [],
                'khan': [],
                'videos': [],
                'audio': [],
                'primary_texts': [{'label': 'Simon Kuznets — Economic Growth and Income Inequality',
                                   'url': 'https://www.nber.org/system/files/chapters/c1584/c1584.pdf',
                                   'era': '1955',
                                   'tradition': 'Empirical',
                                   'note': 'Kuznets curve: inequality rising with early '
                                           'industrialization then falling.'},
                                  {'label': 'Thomas Piketty — Capital in the Twenty-First Century '
                                            '(selected chapters)',
                                   'url': 'https://piketty.pse.ens.fr/files/Piketty2014Capital.pdf',
                                   'era': '2014',
                                   'tradition': 'Distribution/Political Economy',
                                   'note': 'Capital returns exceeding growth drive wealth '
                                           'concentration.'},
                                  {'label': 'Raj Chetty — Behavioral Economics and Public Policy',
                                   'url': 'https://scholar.harvard.edu/files/chetty/files/behavioral_policy.pdf',
                                   'era': '2015',
                                   'tradition': 'Behavioral/Public Finance',
                                   'note': 'Behavioral biases integrated into tax and transfer '
                                           'policy design.'}]},
  'openstax': {'core': [{'chapter': 15,
                         'label': 'OpenStax Ch 15: Poverty and Economic Inequality',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/15-introduction-to-poverty-and-economic-inequality'}],
               'optional': [{'chapter': 12,
                             'label': 'OpenStax Ch 12: Environmental Protection and Negative '
                                      'Externalities',
                             'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/12-introduction-to-environmental-protection-and-negative-externalities'},
                            {'chapter': 13,
                             'label': 'OpenStax Ch 13: Positive Externalities and Public Goods',
                             'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/13-introduction-to-positive-externalities-and-public-goods'},
                            {'chapter': 18,
                             'label': 'OpenStax Ch 18: Public Economy',
                             'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/18-introduction-to-public-economy'}]},
  'format': 'IN_PERSON'},
 {'id': 9,
  'title': 'Firms & Cost of Production',
  'short_desc': 'Firms, economic profit, production functions, principal-agent problems, short-run '
                'costs, and economies of scale.',
  'big_questions': '### Big Questions\n\n- What is the difference between accounting and economic profits?\n- How do the four types of markets shape firm behavior?\n- How do short-run production, short-run costs, economies of scale, and long-run costs fit together?',
  'tier1_definitions': '### Tier 1 — Math Support & Core Definitions\n'
                       '\n'
                       'Tier 1 is the focus for students who may have difficulty with mathematics coming into the course. By the end of Tier 1, students should be able to define:\n'
                       '\n'
                       '- **economic profit**\n'
                       '- **accounting profit**\n'
                       '- **explicit cost**\n'
                       '- **implicit cost**\n'
                       '- **total product**\n'
                       '- **marginal product**\n'
                       '- **fixed cost**\n'
                       '- **variable cost**\n'
                       '- **ATC**\n'
                       '- **AVC**\n'
                       '- **MC**\n'
                       '- **LRAC**',
  'tier2_solid': '### Tier 2 — General Education Standard (Assessment Tier)\n\nTier 2 is the general education target: graphs, tables, formulas, and algebraic reasoning for students taking the course as part of general education. No coding, calculus, or linear algebra is required. By the end of Tier 2, students should be able to:\n\n- **Part 1:** Explore a sample problem to distinguish between accounting and economic profits.\n- **Part 2:** Develop an understanding of the four types of markets and how we will explore firm behavior within them.\n- **Part 3:** Engage with a Manager at a large firm to explore firm structure and decision making.\n- **Part 4:** Develop familiarity with short-run production for a firm, distinguishing between total, average, and marginal product.\n- **Part 5:** Develop familiarity with short-run costs of a firm, distinguishing between total, average and marginal cost.\n- **Part 6:** Apply what you have learned in previous sections to explore economies of scale and differences between long-run and short-run productions and costs.',
  'tier3_extensions': '### Tier 3 — Quantitative Extension\n'
                      '\n'
                      'Tier 3 is a quantitative extension for deeper work: advanced algebra, '
                      'introductory calculus, linear algebra, coding, simulation, or more rigorous '
                      'mathematical proof. By the end of Tier 3, students can choose to:\n'
                      '\n'
                      '- Derive MC and average cost relationships from total cost functions using '
                      'calculus.\n'
                      '- Use advanced algebra to compare short-run and long-run cost envelopes.\n'
                      '- Code cost tables and graphs from production functions.\n'
                      '- Use optimization logic to connect firm boundaries, transaction costs, and '
                      'principal-agent constraints.\n'
                      '- Derive cost curves from production functions and use code to generate '
                      'cost-table sensitivity analysis.',
  'materials': {'models': [{'label': 'Economic vs Accounting Profit',
                            'url': '?model=Economic%20vs%20Accounting%20Profit'},
                           {'label': 'Cost of Production', 'url': '?model=Cost%20of%20Production'},
                           {'label': 'Economies of Scale', 'url': '?model=Economies%20of%20Scale'}],
                'readings': [{'chapter': 7,
                              'label': 'OpenStax Ch 7: Production, Costs, and Industry Structure',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/7-introduction-to-production-costs-and-industry-structure'}],
                'extensions': [],
                'labs': [],
                'khan': [{'label': 'Module 9 Quiz - Production & Costs (Quiz 1)',
                          'url': 'https://www.khanacademy.org/economics-finance-domain/ap-microeconomics/production-cost-and-the-perfect-competition-model-temporary/33-long-run-p/quiz/production-cost-and-the-perfect-competition-model-temporary-quiz-1'},
                         {'label': 'Module 9 Quiz - Production & Costs (Quiz 2)',
                          'url': 'https://www.khanacademy.org/economics-finance-domain/ap-microeconomics/production-cost-and-the-perfect-competition-model-temporary/profit-maximization/quiz/production-cost-and-the-perfect-competition-model-temporary-quiz-2'}],
                'videos': [],
                'audio': [],
                'primary_texts': [{'label': 'Alfred Marshall — Principles (Book III: Demand, '
                                            'Utility, and Value)',
                                   'url': 'https://oll.libertyfund.org/title/marshall-principles-of-economics-8th-ed',
                                   'era': '1890',
                                   'tradition': 'Neoclassical',
                                   'note': 'Marginal utility and demand curves; consumer surplus '
                                           'framing.'},
                                  {'label': 'Jean Tirole — Market Power and Regulation (Nobel '
                                            'Lecture)',
                                   'url': 'https://www.nobelprize.org/uploads/2018/06/tirole-lecture.pdf',
                                   'era': '2014',
                                   'tradition': 'Industrial Organization',
                                   'note': 'Industrial organization theory for market power and '
                                           'regulation incentives.'}],
                'slides': 'https://docs.google.com/presentation/d/1_7e-cwVSdUN-C4VBvpeJj1gVNNsLCcRV/edit?usp=sharing&ouid=106497440738727652653&rtpof=true&sd=true',
                'guided_notes': 'https://www.canva.com/design/DAGkcbojPIk/vHqfUO_zB6h3uWfMkr-bdg/view?utm_content=DAGkcbojPIk&utm_campaign=designshare&utm_medium=link2&utm_source=uniquelinks&utlId=hd509949099'},
  'openstax': {'core': [{'chapter': 7,
                         'label': 'OpenStax Ch 7: Production, Costs, and Industry Structure',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/7-introduction-to-production-costs-and-industry-structure'}],
               'optional': [{'chapter': 8,
                             'label': 'OpenStax Ch 8: Perfect Competition',
                             'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/8-introduction-to-perfect-competition'}]},
  'format': 'IN_PERSON'},
 {'id': 10,
  'title': 'Profit Maximization',
  'short_desc': 'MR = MC, perfect competition, shutdown, long-run entry and exit, monopoly, and '
                'price discrimination.',
  'big_questions': '### Big Questions\n\n- What are the economic and mathematical theories behind profit maximization?\n- How does marginal analysis determine profit-maximizing short-run and long-run production decisions for a perfectly competitive firm?\n- How does monopoly change profit-maximizing behavior and long-run efficiency?',
  'tier1_definitions': '### Tier 1 — Math Support & Core Definitions\n'
                       '\n'
                       'Tier 1 is the focus for students who may have difficulty with mathematics coming into the course. By the end of Tier 1, students should be able to define:\n'
                       '\n'
                       '- **total revenue**\n'
                       '- **marginal revenue**\n'
                       '- **MR = MC**\n'
                       '- **price taker**\n'
                       '- **shutdown rule**\n'
                       '- **break-even**\n'
                       '- **entry**\n'
                       '- **exit**\n'
                       '- **monopoly**\n'
                       '- **barriers to entry**\n'
                       '- **price discrimination**',
  'tier2_solid': '### Tier 2 — General Education Standard (Assessment Tier)\n\nTier 2 is the general education target: graphs, tables, formulas, and algebraic reasoning for students taking the course as part of general education. No coding, calculus, or linear algebra is required. By the end of Tier 2, students should be able to:\n\n- **Part 1:** Review Unit 9\n- **Part 2:** Develop an understanding of the economic and mathematical theories behind profit maximization.\n- **Part 3:** Define rules and use marginal analysis to determine profit-maximizing, short-run production decisions for a perfectly competitive firm.\n- **Part 4:** Define rules and use marginal analysis to determine profit maximizing, long-run production decisions for a perfectly competitive firm.\n- **Part 5:** Define rules and use marginal analysis to determine profit-maximizing, short-run production decisions for a monopoly.\n- **Part 6:** Apply what you have learned to compare the efficiency of perfect competition and monopoly markets in the long-run.',
  'tier3_extensions': '### Tier 3 — Quantitative Extension\n'
                      '\n'
                      'Tier 3 is a quantitative extension for deeper work: advanced algebra, '
                      'introductory calculus, linear algebra, coding, simulation, or more rigorous '
                      'mathematical proof. By the end of Tier 3, students can choose to:\n'
                      '\n'
                      '- Derive MR from inverse demand and solve MR = MC symbolically.\n'
                      '- Use calculus to compare profit maximization under perfect competition and '
                      'monopoly.\n'
                      '- Code profit surfaces, shutdown regions, and entry-exit simulations.\n'
                      '- Use linear systems to compare monopoly, competition, and '
                      'price-discrimination outcomes.\n'
                      '- Use calculus or coded simulations to compare competitive, monopoly, '
                      'shutdown, and price-discrimination outcomes.',
  'materials': {'models': [{'label': 'Perfect Competition: Profit Maximization',
                            'url': '?model=Perfect%20Competition:%20Profit%20Maximization'},
                           {'label': 'Perfect Competition: Shutdown Point',
                            'url': '?model=Perfect%20Competition:%20Shutdown%20Point'},
                           {'label': 'Long-Run Equilibrium and Firm Exit',
                            'url': '?model=Long-Run%20Equilibrium%20and%20Firm%20Exit'},
                           {'label': 'Monopoly and Monopolistic Competition',
                            'url': '?model=Monopoly%20and%20Monopolistic%20Competition'},
                           {'label': 'Price Discrimination',
                            'url': '?model=Price%20Discrimination'}],
                'readings': [{'chapter': 8,
                              'label': 'OpenStax Ch 8: Perfect Competition',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/8-introduction-to-perfect-competition'},
                             {'chapter': 9,
                              'label': 'OpenStax Ch 9: Monopoly',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/9-introduction-to-a-monopoly'}],
                'extensions': [],
                'labs': [],
                'khan': [{'label': 'Module 10 Quiz - Production & Costs (Quiz 2)',
                          'url': 'https://www.khanacademy.org/economics-finance-domain/ap-microeconomics/production-cost-and-the-perfect-competition-model-temporary/profit-maximization/quiz/production-cost-and-the-perfect-competition-model-temporary-quiz-2'},
                         {'label': 'Module 10 Quiz - Imperfect Competition (Quiz 1)',
                          'url': 'https://www.khanacademy.org/economics-finance-domain/ap-microeconomics/imperfect-competition/ap-price-discrimination-tutorial/quiz/imperfect-competition-quiz-1'},
                         {'label': 'Module 10 Quiz - Production & Costs (Quiz 3)',
                          'url': 'https://www.khanacademy.org/economics-finance-domain/ap-microeconomics/production-cost-and-the-perfect-competition-model-temporary/ap-perfect-competition/quiz/production-cost-and-the-perfect-competition-model-temporary-quiz-3'}],
                'videos': [],
                'audio': [],
                'primary_texts': [{'label': 'Alfred Marshall — Principles (Book III: Demand, '
                                            'Utility, and Value)',
                                   'url': 'https://oll.libertyfund.org/title/marshall-principles-of-economics-8th-ed',
                                   'era': '1890',
                                   'tradition': 'Neoclassical',
                                   'note': 'Marginal utility and demand curves; consumer surplus '
                                           'framing.'},
                                  {'label': 'Jean Tirole — Market Power and Regulation (Nobel '
                                            'Lecture)',
                                   'url': 'https://www.nobelprize.org/uploads/2018/06/tirole-lecture.pdf',
                                   'era': '2014',
                                   'tradition': 'Industrial Organization',
                                   'note': 'Industrial organization theory for market power and '
                                           'regulation incentives.'}],
                'slides': 'https://docs.google.com/presentation/d/1kozm3rG9vJ-KhPnAL-GDfC06sFFVNAfX/edit?usp=sharing&ouid=106497440738727652653&rtpof=true&sd=true',
                'guided_notes': 'https://www.canva.com/design/DAGkcYB26Qo/lLa38_j6olYROBl_PJURnw/view?utm_content=DAGkcYB26Qo&utm_campaign=designshare&utm_medium=link2&utm_source=uniquelinks&utlId=h3f3941310e'},
  'openstax': {'core': [{'chapter': 8,
                         'label': 'OpenStax Ch 8: Perfect Competition',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/8-introduction-to-perfect-competition'},
                        {'chapter': 9,
                         'label': 'OpenStax Ch 9: Monopoly',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/9-introduction-to-a-monopoly'}],
               'optional': [{'chapter': 10,
                             'label': 'OpenStax Ch 10: Monopolistic Competition and Oligopoly',
                             'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/10-introduction-to-monopolistic-competition-and-oligopoly'}]},
  'format': 'IN_PERSON'},
 {'id': 11,
  'title': 'Imperfect Competition & Game Theory',
  'short_desc': 'Monopolistic competition, oligopoly, payoff matrices, Nash equilibrium, prison '
                'dilemmas, and antitrust.',
  'big_questions': '### Big Questions\n\n- How does game theory explain profit maximizing behavior for oligopolies?\n- How does monopolistic competition deepen understanding of the spectrum of market competitiveness?\n- How does market regulation through antitrust law respond to imperfect competition?',
  'tier1_definitions': '### Tier 1 — Math Support & Core Definitions\n'
                       '\n'
                       'Tier 1 is the focus for students who may have difficulty with mathematics coming into the course. By the end of Tier 1, students should be able to define:\n'
                       '\n'
                       '- **monopolistic competition**\n'
                       '- **game theory**\n'
                       '- **players**\n'
                       '- **strategies**\n'
                       '- **payoffs**\n'
                       '- **dominant strategy**\n'
                       '- **Nash equilibrium**\n'
                       '- **Pareto efficiency**\n'
                       '- **oligopoly**\n'
                       '- **cartel**\n'
                       '- **HHI**',
  'tier2_solid': '### Tier 2 — General Education Standard (Assessment Tier)\n\nTier 2 is the general education target: graphs, tables, formulas, and algebraic reasoning for students taking the course as part of general education. No coding, calculus, or linear algebra is required. By the end of Tier 2, students should be able to:\n\n- **Part 1:** Learn the basic framework of game theory and play the prisoner’s dilemma game.\n- **Part 2:** Use game theory to explore profit maximizing behavior for oligopolies.\n- **Part 3:** Explore profit maximizing behavior for monopolistic competition in order to deepen your understanding of the spectrum of market competitiveness.\n- **Part 4:** Develop a brief understanding of market regulation through antitrust law.\n- **Part 5:** Refine your understanding of market models by modeling three distinct markets in business centers.',
  'tier3_extensions': '### Tier 3 — Quantitative Extension\n'
                      '\n'
                      'Tier 3 is a quantitative extension for deeper work: advanced algebra, '
                      'introductory calculus, linear algebra, coding, simulation, or more rigorous '
                      'mathematical proof. By the end of Tier 3, students can choose to:\n'
                      '\n'
                      '- Represent games with matrices and solve for Nash equilibria '
                      'systematically.\n'
                      '- Use linear algebra notation for mixed strategies as an optional '
                      'extension.\n'
                      "- Code repeated prisoner's dilemma, cartel-cheating, or merger "
                      'simulations.\n'
                      '- Extend oligopoly models with Cournot or Bertrand algebra.\n'
                      '- Use matrix notation, mixed-strategy logic, or repeated-game simulations '
                      'to extend the game theory models.',
  'materials': {'models': [{'label': 'Monopolistic Competition and Oligopoly',
                            'url': '?model=Monopolistic%20Competition%20and%20Oligopoly'},
                           {'label': 'Game Theory', 'url': '?model=Game%20Theory'},
                           {'label': 'Antitrust HHI and Merger Analysis',
                            'url': '?model=Antitrust%20HHI%20and%20Merger%20Analysis'},
                           {'label': 'Competition, Information, and Fairness',
                            'url': '?model=Competition,%20Information,%20and%20Fairness'}],
                'readings': [{'chapter': 8,
                              'label': 'OpenStax Ch 8: Perfect Competition',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/8-introduction-to-perfect-competition'},
                             {'chapter': 9,
                              'label': 'OpenStax Ch 9: Monopoly',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/9-introduction-to-a-monopoly'},
                             {'chapter': 10,
                              'label': 'OpenStax Ch 10: Monopolistic Competition and Oligopoly',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/10-introduction-to-monopolistic-competition-and-oligopoly'}],
                'extensions': [],
                'labs': [],
                'khan': [{'label': 'Module 11 Quiz - Imperfect Competition (Quiz 1)',
                          'url': 'https://www.khanacademy.org/economics-finance-domain/ap-microeconomics/imperfect-competition/ap-price-discrimination-tutorial/quiz/imperfect-competition-quiz-1'},
                         {'label': 'Module 11 Quiz - Imperfect Competition (Quiz 2)',
                          'url': 'https://www.khanacademy.org/economics-finance-domain/ap-microeconomics/imperfect-competition/oligopoly-and-game-theory/quiz/imperfect-competition-quiz-2'}],
                'videos': [],
                'audio': [],
                'primary_texts': [{'label': 'George Akerlof — The Market for Lemons',
                                   'url': 'https://www2.bc.edu/thomas-chemmanur/phdfincorp/MF891%20papers/Akerlof%201970.pdf',
                                   'era': '1970',
                                   'tradition': 'Information Economics',
                                   'note': 'Asymmetric information and adverse selection can '
                                           'unravel markets.'},
                                  {'label': 'Michael Spence — Job Market Signaling',
                                   'url': 'https://web.stanford.edu/~sbale/Spence1973.pdf',
                                   'era': '1973',
                                   'tradition': 'Information Economics',
                                   'note': 'Costly signaling to sort worker ability in job '
                                           'markets.'},
                                  {'label': 'Jean Tirole — Market Power and Regulation (Nobel '
                                            'Lecture)',
                                   'url': 'https://www.nobelprize.org/uploads/2018/06/tirole-lecture.pdf',
                                   'era': '2014',
                                   'tradition': 'Industrial Organization',
                                   'note': 'Industrial organization theory for market power and '
                                           'regulation incentives.'}],
                'slides': 'https://docs.google.com/presentation/d/17glhtdEYx9r_YZ3JdtStAWopq3TeAH8i/edit?usp=sharing&ouid=106497440738727652653&rtpof=true&sd=true',
                'guided_notes': 'https://www.canva.com/design/DAGktlosX4M/H1KvxNqI7QmVFe42oYD41w/view?utm_content=DAGktlosX4M&utm_campaign=designshare&utm_medium=link2&utm_source=uniquelinks&utlId=h5b758b3187'},
  'openstax': {'core': [{'chapter': 8,
                         'label': 'OpenStax Ch 8: Perfect Competition',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/8-introduction-to-perfect-competition'},
                        {'chapter': 9,
                         'label': 'OpenStax Ch 9: Monopoly',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/9-introduction-to-a-monopoly'},
                        {'chapter': 10,
                         'label': 'OpenStax Ch 10: Monopolistic Competition and Oligopoly',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/10-introduction-to-monopolistic-competition-and-oligopoly'}],
               'optional': [{'chapter': 11,
                             'label': 'OpenStax Ch 11: Monopoly and Antitrust Policy',
                             'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/11-introduction-to-monopoly-and-antitrust-policy'},
                            {'chapter': 16,
                             'label': 'OpenStax Ch 16: Information, Risk, and Insurance',
                             'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/16-introduction-to-information-risk-and-insurance'}]},
  'format': 'IN_PERSON'},
 {'id': 12,
  'title': 'Policy, Paradox & Human Perspectives',
  'short_desc': 'Types of goods, externalities, Pigouvian policy, dynamic policy design, '
                'behavioral policy, public goods games, and course synthesis.',
  'big_questions': '### Big Questions\n\n- Why is the market quantity less than the socially optimal quantity?\n- How might education subsidies help achieve the socially optimal outcome?\n- What size subsidy would be needed to reach the social optimum?\n- What are some other real-world examples of positive externalities?',
  'tier1_definitions': '### Tier 1 — Math Support & Core Definitions\n'
                       '\n'
                       'Tier 1 is the focus for students who may have difficulty with mathematics coming into the course. By the end of Tier 1, students should be able to define:\n'
                       '\n'
                       '- **rivalry**\n'
                       '- **excludability**\n'
                       '- **public goods**\n'
                       '- **common resources**\n'
                       '- **free rider**\n'
                       '- **positive externality**\n'
                       '- **negative externality**\n'
                       '- **Pigouvian tax**\n'
                       '- **Pigouvian subsidy**\n'
                       '- **cap-and-trade**\n'
                       '- **nudge**\n'
                       '- **ultimatum game**',
  'tier2_solid': '### Tier 2 — General Education Standard (Assessment Tier)\n\nTier 2 is the general education target: graphs, tables, formulas, and algebraic reasoning for students taking the course as part of general education. No coding, calculus, or linear algebra is required. By the end of Tier 2, students should be able to:\n\n- **Part 0**: Review Unit 4 and Game Theory\n- **Part 1:** Determine the difference between private goods, public goods, club goods, and common resources.\n- **Part 2:** Develop a model for positive and negative externalities.\n- **Part 3:** Model how governments respond to positive and negative externalities in private markets with taxes and subsidies.\n- **Part 4:** Use game theory to explore causal policy under real-world circumstances.\n- **Part 5:** Explore the cost-benefit analysis behind the decision to instal a traffic light.\n- **Part 6:** Apply what you have learned to a real-world policy simulation.\n- **Activity A:** Learn about Positive Externalities\n- **Activity B:** Research an Education Market\n- **Activity C:** Write a Policy Memo',
  'tier3_extensions': '### Tier 3 — Quantitative Extension\n'
                      '\n'
                      'Tier 3 is a quantitative extension for deeper work: advanced algebra, '
                      'introductory calculus, linear algebra, coding, simulation, or more rigorous '
                      'mathematical proof. By the end of Tier 3, students can choose to:\n'
                      '\n'
                      '- Use calculus or integration to compute welfare areas under nonlinear '
                      'externality curves.\n'
                      '- Code Pigouvian tax, subsidy, cap-and-trade, and public-goods game '
                      'simulations.\n'
                      '- Use dynamic equations to represent path dependence, adjustment costs, and '
                      'policy timing.\n'
                      '- Apply rigorous mathematical and computational critique to what policy '
                      'models can and cannot see.\n'
                      '- Use dynamic models or simulations to compare Pigouvian policy, '
                      'cap-and-trade, public provision, and behavioral nudges over time.',
  'materials': {'models': [{'label': 'Types of Goods', 'url': '?model=Types%20of%20Goods'},
                           {'label': 'Public Goods and Common Resources',
                            'url': '?model=Public%20Goods%20and%20Common%20Resources'},
                           {'label': 'Externalities and Pigouvian Policy',
                            'url': '?model=Externalities%20and%20Pigouvian%20Policy'},
                           {'label': 'Tax Incidence', 'url': '?model=Tax%20Incidence'},
                           {'label': 'Game Theory', 'url': '?model=Game%20Theory'},
                           {'label': 'Behavioral Policy', 'url': '?model=Behavioral%20Policy'},
                           {'label': 'GDP and Wellbeing Limits',
                            'url': '?model=GDP%20and%20Wellbeing%20Limits'}],
                'readings': [{'chapter': 12,
                              'label': 'OpenStax Ch 12: Environmental Protection and Negative '
                                       'Externalities',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/12-introduction-to-environmental-protection-and-negative-externalities'},
                             {'chapter': 13,
                              'label': 'OpenStax Ch 13: Positive Externalities and Public Goods',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/13-introduction-to-positive-externalities-and-public-goods'},
                             {'chapter': 18,
                              'label': 'OpenStax Ch 18: Public Economy',
                              'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/18-introduction-to-public-economy'}],
                'extensions': [],
                'labs': [],
                'khan': [{'label': 'Module 12 Quiz - Market Failure (Quiz 1)',
                          'url': 'https://www.khanacademy.org/economics-finance-domain/ap-microeconomics/ap-consumer-producer-surplus/public-and-private-goods/quiz/ap-consumer-producer-surplus-quiz-1?referrer=upsell'}],
                'videos': [],
                'audio': [],
                'primary_texts': [{'label': 'A.C. Pigou — The Economics of Welfare (Externalities)',
                                   'url': 'https://socialsciences.mcmaster.ca/econ/ugcm/3ll3/pigou/welfare1.pdf',
                                   'era': '1920',
                                   'tradition': 'Welfare',
                                   'note': 'Externalities and corrective taxes/subsidies to align '
                                           'private and social cost.'},
                                  {'label': 'R.H. Coase — The Problem of Social Cost',
                                   'url': 'https://chicagounbound.uchicago.edu/cgi/viewcontent.cgi?article=1002&context=law_and_economics',
                                   'era': '1960',
                                   'tradition': 'Institutional',
                                   'note': 'Transaction costs and property rights; bargaining can '
                                           'solve externalities.'},
                                  {'label': 'Amartya Sen — Equality of What?',
                                   'url': 'https://dash.harvard.edu/bitstream/handle/1/9407592/EqualityofWhat.pdf',
                                   'era': '1979',
                                   'tradition': 'Welfare/Capabilities',
                                   'note': 'Capabilities approach to welfare beyond income; '
                                           'inequality as deprivation.'},
                                  {'label': 'Raj Chetty — Behavioral Economics and Public Policy',
                                   'url': 'https://scholar.harvard.edu/files/chetty/files/behavioral_policy.pdf',
                                   'era': '2015',
                                   'tradition': 'Behavioral/Public Finance',
                                   'note': 'Behavioral biases integrated into tax and transfer '
                                           'policy design.'}],
                'slides': 'https://www.canva.com/design/DAG5q19uVic/8m5nIPLoKl01XmS99hJReA/view?utm_content=DAG5q19uVic&utm_campaign=designshare&utm_medium=link2&utm_source=uniquelinks&utlId=hc51ff648b0',
                'guided_notes': 'https://www.canva.com/design/DAGj-XcXXCQ/HqFSA-Y9C1A6335ZWD98EQ/view?utm_content=DAGj-XcXXCQ&utm_campaign=designshare&utm_medium=link2&utm_source=uniquelinks&utlId=h26cb36b088'},
  'openstax': {'core': [{'chapter': 12,
                         'label': 'OpenStax Ch 12: Environmental Protection and Negative '
                                  'Externalities',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/12-introduction-to-environmental-protection-and-negative-externalities'},
                        {'chapter': 13,
                         'label': 'OpenStax Ch 13: Positive Externalities and Public Goods',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/13-introduction-to-positive-externalities-and-public-goods'},
                        {'chapter': 18,
                         'label': 'OpenStax Ch 18: Public Economy',
                         'url': 'https://openstax.org/books/principles-microeconomics-3e/pages/18-introduction-to-public-economy'}]},
  'format': 'IN_PERSON'}]

# --- Link OpenStax chapters to modules (core + optional) ---
def _build_openstax_item(chapter_num: int):
    chapter = OPENSTAX_CHAPTERS.get(chapter_num)
    if not chapter:
        return None
    return {
        "chapter": chapter_num,
        "label": f"OpenStax Ch {chapter_num}: {chapter['title']}",
        "url": chapter["url"],
    }


_PLAN_BY_MODULE = {m["module"]: m for m in MODULE_OPENSTAX_MAP}

for module in MICRO_MODULES:
    mapping = _PLAN_BY_MODULE.get(module.get("id"))
    if not mapping:
        continue

    core_items = []
    optional_items = []

    for num in mapping.get("core_openstax", []):
        item = _build_openstax_item(num)
        if item:
            core_items.append(item)

    for num in mapping.get("optional_openstax", []):
        item = _build_openstax_item(num)
        if item:
            optional_items.append(item)

    if not core_items and not optional_items:
        continue

    module.setdefault("openstax", {})
    if core_items:
        module["openstax"]["core"] = core_items
    if optional_items:
        module["openstax"]["optional"] = optional_items

    # Put core OpenStax links into required readings (avoid duplicates)
    materials = module.setdefault("materials", {})
    existing_readings = materials.get("readings", [])
    seen_urls = {item["url"] for item in core_items}
    deduped_existing = [r for r in existing_readings if r.get("url") not in seen_urls]
    materials["readings"] = core_items + deduped_existing


# --- Link Primary/Canonical texts (classic + contemporary) to modules ---
def _build_canonical_item(key: str):
    item = CANONICAL_TEXTS.get(key)
    if not item:
        return None
    return {
        "label": item["label"],
        "url": item["url"],
        "era": item["era"],
        "tradition": item["tradition"],
        "note": item.get("note", ""),
    }


for module in MICRO_MODULES:
    keys = MODULE_CANONICAL_MAP.get(module.get("id"), [])
    if not keys:
        continue
    new_items = []
    for k in keys:
        built = _build_canonical_item(k)
        if built:
            new_items.append(built)
    if not new_items:
        continue

    materials = module.setdefault("materials", {})
    existing = materials.get("primary_texts", [])
    existing_urls = {itm.get("url") for itm in existing}
    merged = existing + [itm for itm in new_items if itm["url"] not in existing_urls]
    materials["primary_texts"] = merged


# --- Ingest course links workbook/CSV to fill module resources ---
COURSE_LINKS_FILENAME = "Course Links & Resources 6d6e76ec5670407fb399d4ec39993f2c_all.csv"
COURSE_LINKS_UPDATED_FILENAME = "Course Links-Table 1.csv"
COURSE_LINKS_WORKBOOK = "ECN101_Course_Links_Resources_Updated.xlsx"
_BASE_DIR = Path(__file__).resolve().parent
_RUNNING_MATERIALS_DIR = _BASE_DIR / "data" / "Running Materials"
FINAL_BUILD_DIR = _RUNNING_MATERIALS_DIR / "Final Build"
_COURSE_LINKS_CANDIDATES = [
    _RUNNING_MATERIALS_DIR / COURSE_LINKS_UPDATED_FILENAME, # current updated syllabus source
    _RUNNING_MATERIALS_DIR / COURSE_LINKS_WORKBOOK,
    _BASE_DIR / "data" / COURSE_LINKS_FILENAME,          # new canonical location
    _BASE_DIR / "dev_materials" / COURSE_LINKS_FILENAME, # legacy location (fallback)
]

COURSE_LINKS_PATH = next((p for p in _COURSE_LINKS_CANDIDATES if p.exists()), _RUNNING_MATERIALS_DIR / COURSE_LINKS_UPDATED_FILENAME)


def _parse_module_id(module_str: str):
    match = re.search(r"(\d+)", module_str or "")
    return int(match.group(1)) if match else None


def _merge_unique(existing, new):
    merged = list(existing or [])
    seen = {item.get("url") for item in merged if isinstance(item, dict)}
    for item in new or []:
        url = item.get("url")
        if url and url not in seen:
            merged.append(item)
            seen.add(url)
    return merged


def _load_workbook_rows(workbook_path: Path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["Course Links"] if "Course Links" in workbook.sheetnames else workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    records = []
    for row in rows:
        record = {
            headers[idx]: value
            for idx, value in enumerate(row)
            if idx < len(headers) and headers[idx]
        }
        records.append(record)
    return records


def _load_csv_rows(csv_path: Path):
    with csv_path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _load_course_links(source_path: Path):
    if not source_path.exists():
        return {}

    legacy_module_remap = {
        2: 3,  # old Supply & Demand -> Supply and Demand
        3: 4,  # old Elasticity -> Market Analysis
        4: 4,  # old Welfare & Intervention -> Market Analysis
        5: 5,  # Factors remains Module 5 in the final outline
        6: 2,  # old Choice -> Choice
        7: 6,  # old Capitalism -> Bridge
        8: 7,  # old Inequality -> Structural Inequality Core
    }

    by_module = {}
    rows = _load_workbook_rows(source_path) if source_path.suffix.lower() == ".xlsx" else _load_csv_rows(source_path)

    for row in rows:
        module_id = row.get("New Module #")
        if module_id is None:
            module_id = _parse_module_id(row.get("Module", ""))
            if module_id:
                module_id = legacy_module_remap.get(module_id, module_id)
        else:
            module_id = int(module_id)
        if not module_id:
            continue

        url = str(row.get("Document URL") or row.get("document url") or "").strip()
        resource_type = str(row.get("Resource Type") or row.get("Resource Type 1") or "").strip().lower()
        title = str(row.get("Link Title") or row.get("\ufeffLink Title") or "").strip()
        desc = str(row.get("Description") or "").strip()
        module_title = str(row.get("New Module Title") or row.get("Module Title") or "").strip()
        label = title or desc or module_title or "Resource"

        if not url or not resource_type:
            continue

        record = by_module.setdefault(
            module_id,
            {"slides": [], "guided_notes": [], "quizzes": [], "required_readings": [], "optional_readings": []},
        )

        if label == module_title and resource_type in {"lecture slides", "guided notes"}:
            label = f"{module_title} {'Slides' if resource_type == 'lecture slides' else 'Guided Notes'}"

        item = {"label": label, "url": url}
        if resource_type == "lecture slides":
            record["slides"].append(item)
        elif resource_type == "guided notes":
            record["guided_notes"].append(item)
        elif resource_type in {"worksheet", "worksheets", "activity", "activities"}:
            record.setdefault("labs", []).append(item)
        elif resource_type == "quiz":
            record["quizzes"].append(item)
        elif resource_type == "required reading - textbook":
            record["required_readings"].append(item)
        elif resource_type == "optional reading - textbook":
            record["optional_readings"].append(item)

    return by_module


COURSE_LINKS_BY_MODULE = _load_course_links(COURSE_LINKS_PATH)

for module in MICRO_MODULES:
    links = COURSE_LINKS_BY_MODULE.get(module.get("id"))
    if not links:
        continue

    materials = module.setdefault("materials", {})

    if links.get("slides"):
        materials["slides"] = links["slides"]
    if links.get("guided_notes"):
        materials["guided_notes"] = links["guided_notes"]
    if links.get("labs"):
        materials["labs"] = links["labs"]
    materials["khan"] = links.get("quizzes", [])

    materials["readings"] = links.get("required_readings", [])

    openstax_links = module.setdefault("openstax", {})
    openstax_links["optional"] = links.get("optional_readings", [])


def _final_build_files(final_build_dir: Path):
    files_by_module = {}
    if not final_build_dir.exists():
        return files_by_module

    for path in sorted(final_build_dir.iterdir()):
        if path.suffix.lower() not in {".docx", ".pptx"}:
            continue
        module_id = _parse_module_id(path.name)
        if not module_id:
            continue
        key = "worksheet_files" if path.suffix.lower() == ".docx" else "slide_files"
        label = path.stem.replace("_", " ")
        files_by_module.setdefault(module_id, {}).setdefault(key, []).append(
            {"label": label, "path": str(path)}
        )
    return files_by_module


FINAL_BUILD_FILES_BY_MODULE = _final_build_files(FINAL_BUILD_DIR)

for module in MICRO_MODULES:
    final_files = FINAL_BUILD_FILES_BY_MODULE.get(module.get("id"))
    if not final_files:
        continue
    materials = module.setdefault("materials", {})
    if final_files.get("slide_files"):
        materials["slide_files"] = final_files["slide_files"]
    if final_files.get("worksheet_files"):
        materials["worksheet_files"] = final_files["worksheet_files"]


from data.module_big_questions import format_big_questions
from data.module_summaries import MODULE_SUMMARIES

for module in MICRO_MODULES:
    final_big_questions = format_big_questions(module.get("id"))
    if final_big_questions:
        module["big_questions"] = final_big_questions

    module_summary = MODULE_SUMMARIES.get(module.get("id"))
    if not module_summary:
        module.pop("overview_intuition", None)
        continue

    vocabulary = ", ".join(f"`{term}`" for term in module_summary["vocabulary"])
    module["overview_intuition"] = (
        f"### Module {module['id']} Overview\n"
        f"{module_summary['summary']}\n\n"
        f"**History and research connection:** {module_summary['history_research']}\n\n"
        f"**Key vocabulary:** {vocabulary}"
    )
